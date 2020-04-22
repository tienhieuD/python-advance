import os
import re
from io import BytesIO

import openpyxl
from xlsxwriter.utility import xl_rowcol_to_cell
from jinja2 import Template
from openpyxl import utils
from openpyxl.writer.excel import save_virtual_workbook


def read_xlsx():
    break_line = '@break-line@'
    col = '@col@'
    jj = 'jj:'

    file = open('C:\\Users\\hieudt\\Desktop\\python-advance\\excel_template\\template_2.xlsx', 'rb')
    b = BytesIO(file.read())
    print(b)
    wb = openpyxl.load_workbook(filename=b)
    # wb = openpyxl.load_workbook(filename='template_2.xlsx')
    main_sheet = wb.active

    values = []
    for row in main_sheet.rows:
        row_vals = []
        for cell in row:
            row_vals.append(cell.value)
        values.append(row_vals)
    print(values)

    strs = ""
    for val in values:
        strs += col.join("%s<<%s>>" % (v, type(v).__name__) for v in val)+break_line
    tmpl = Template(strs)
    res = tmpl.render()

    rows = res.split(break_line)
    datas = []
    for r in rows:
        row = r.split(col)
        # if jj in row[0]:
        #     continue
        for cell in row:
            if cell == 'None:':
                cell = None
        datas.append(row)
    print(datas)

    for row in range(len(datas)):
        for col in range(len(datas[row])):
            col_letter = utils.get_column_letter(col+1)
            cell = "%s%s" % (col_letter, row+1)
            value = datas[row][col]
            if re.search(r"<<(.+?)>>", value):
                value_type = re.search(r"<<(.+?)>>", value).group(1)
                value = re.sub(r'<<(.+?)>>', '', value)
                finval = None if value_type == 'NoneType' else eval(value_type)(value)
                main_sheet[cell] = finval

    x = BytesIO(save_virtual_workbook(wb))
    wb.save('export.xlsx')
    os.system('export.xlsx')
    print(x)
    print(save_virtual_workbook(wb))


def get_simple_xlsx(template_filename=None, begin_data_cell='A1', data_array=None):
    """
    :param template_filename: path or file-like object (BytesIO)
    :param begin_data_cell: cell to start writing data, keep format to following cells
    :param data_array: data array
    :return: file-like object (BytesIO)
    """
    workbook = openpyxl.load_workbook(filename=template_filename)
    main_sheet = workbook.active

    row_length = len(data_array)
    col_length = len(data_array and data_array[0] or [])

    repatt = re.search(r'^(\w+?)(\d+?)', begin_data_cell)
    begin_col_index = utils.column_index_from_string(repatt.group(1)) - 1
    begin_row_index = int(repatt.group(2)) - 1

    main_sheet.insert_rows(begin_row_index + 2, row_length)
    for row_index in range(row_length):
        for col_index in range(col_length):
            # Value assign
            value = data_array[row_index][col_index]
            cell_to_write = xl_rowcol_to_cell(
                row_index + begin_row_index,
                col_index + begin_col_index)
            main_sheet[cell_to_write].value = value
            # Style assign
            _style = main_sheet[begin_row_index + 1][col_index]._style
            main_sheet[cell_to_write]._style = _style

    workbook.save('export.xlsx')
    return BytesIO(save_virtual_workbook(workbook))


if __name__ == '__main__':
    # read_xlsx()
    get_simple_xlsx(
        template_filename='C:\\Users\\hieudt\\Desktop\\python-advance\\excel_template\\template_2.xlsx',
        begin_data_cell='A3',
        data_array=[
            [1, 2, 3, 4, 5, 6, 7, 8, 9, 0],
            [1, 2, 3, 4, 5, 6, 7, 8, 9, 0],
            [1, 2, 3, 4, 5, 6, 7, 8, 9, 0],
            [1, 2, 3, 4, 5, 6, 7, 8, 9, 0],
            [1, 2, 3, 4, 5, 6, 7, 8, 9, 0],
            [1, 2, 3, 4, 5, 6, 7, 8, 9, 0],
            [1, 2, 3, 4, 5, 6, 7, 8, 9, 0],
            [1, 2, 3, 4, 5, 6, 7, 8, 9, 0],
        ],
    )